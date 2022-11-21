from __future__ import nested_scopes
from email import message
from xml.dom.domreg import registered
from flask import Flask, request, session, render_template, redirect, jsonify
from flask_session import Session
from datetime import timedelta
from flask_mobility import Mobility
import atexit
import networker
import pickle
import jdatetime
from os.path import exists
from openpyxl import load_workbook

STUDENTS_LIST_FILE = "students-list.xlsx"
AP_SSID = 'Kian Attendence app'
AP_PASS = '12345678'

def load_list_file():
    wb = load_workbook(STUDENTS_LIST_FILE)
    return wb,wb.active

_,ws = load_list_file()
std_nums = [str(row[0].value) for row in ws.iter_rows(min_row=2,max_col=1) if row[0].value is not None]
dates = [col[0].value for col in ws.iter_cols(3,None,1,1) if col[0].value is not None]

app = Flask(__name__,static_folder=r'templates\assets')
app.config['SESSION_PERMENENT'] = True
app.config['SESSION_TYPE'] = 'filesystem'
app.permanent_session_lifetime = timedelta(minutes=30)
Session(app)
Mobility(app)

def mac2std(mac:str)->str:
    if exists('mac-std.dict'):
        mac_std:dict = pickle.load(open('mac-std.dict','rb'))
        return mac_std.get(mac)
    else:
        return None

def save_mac_std(mac:str, std_num:str):
    mac_std=dict()
    if exists('mac-std.dict'):
        with open('mac-std.dict','r+b') as f:
            mac_std= pickle.load(f)
            mac_std[mac] = std_num
            pickle.dump(mac_std,f)
    else:
        mac_std[mac] = std_num
        pickle.dump(mac_std,open('mac-std.dict','wb'))

@app.before_request
def before_req():
    if request.remote_addr == '127.0.0.1':
        session['mac'] = 'local-client'
        session['std_num'] = mac2std(session['mac'])
    elif not request.MOBILE:
        return '<h1>Please use your mobile-phone, other devices are blocked 😉</h1>'
    if session.get('mac') is None:
        session['mac'] = networker.get_mac(request.remote_addr)
        session['std_num'] = mac2std(session['mac'])

@app.route('/')
def index():
    if session['std_num'] is None:
        return render_template('index.html',registered=False)
    else:
        std_num = session['std_num']
        # Registration of student attendance
        index = std_nums.index(std_num)+2
        wb,ws=load_list_file()
        ws.cell(index,column=ws.max_column).value='P'
        wb.save(STUDENTS_LIST_FILE)
        num,name,*attendence = map(lambda x:str(x.value), ws[index])
        print(f'{num=}, {attendence=}, {dates=}')
        return render_template('index.html',registered=True,student_name=name,attendence=attendence,dates=dates)

@app.route('/api/v1/register')
def register_v1():
    std_num = request.args.get('std_num')
    if (session['std_num'] is not None or std_num is None) and session['mac'] !='local-client':  # if user registred before or no std_num sent return error
        return jsonify(done=False, error=1)  # error 0: registration faild
    
    # register student
    print(f"registering user mac:{session['mac']}, std_num:\"{std_num}\"")
    if std_num in std_nums:
        print('registered')
        save_mac_std(session['mac'], std_num)
        return jsonify(done=True)
    else:
        print('std_num is not in the list')
        return jsonify(done=False, error=0)  # error 1: std_num is not validated

@app.route('/api/v1/attendence')
def attendence_v1():
    std_num = session['std_num']
    if std_num is None:
        return jsonify(done=False, error=1)  # error 0: you mast do registration
    
    if not std_num in std_nums:
        return jsonify(done=False, error=0) # error 1: std_num is not validated
    
    # Registration of student attendance
    row = std_nums.index(std_num)+2
    wb,ws=load_list_file()
    ws.cell(row,column=max(3,ws.max_column)).value='P'
    wb.save(STUDENTS_LIST_FILE)
    num,name,*attendence = map(lambda x:str(x.value), ws[row])
    print(f'{num=},{name=} {attendence=}, {dates=}')
    content = render_template('attendence.html',attendence=attendence,dates=dates)
    return jsonify(done=True,name=name, content=content)
    
def on_exit():
    networker.stop_ap()

atexit.register(on_exit)

if __name__=='__main__':
    wb,ws = load_list_file()
    date = '{d.month}/{d.day}'.format(d=jdatetime.date.today())
    for col in ws.iter_cols(max_row=1,min_col=3):
        if date==col[0].value:
            break
    else:
        new_col = ws.max_column+1
        ws.cell(row=1,column=new_col).value=date
        for row in ws.iter_rows(min_row=2,min_col=new_col,max_col=new_col):
            row[0].value='A'
        wb.save(STUDENTS_LIST_FILE)

    networker.config_softap(AP_SSID,AP_PASS)
    networker.start_ap()
    app.run('0.0.0.0',80,True)