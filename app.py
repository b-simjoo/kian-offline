from flask import Flask, request, session, render_template, redirect, jsonify, g, url_for, abort
from flask_session import Session
from flask_mobility import Mobility
from flask_expects_json import expects_json
from datetime import timedelta, datetime
from openpyxl import load_workbook
from getmac import get_mac_address
from tempfile import NamedTemporaryFile
from peewee import DoesNotExist
from jsonschema import ValidationError

from model import db, Student, Device, Attendance, Score, Meeting, _TABLES_
from customjsonencoder import CustomJSONEncoder
from schema import LOGIN_SCHEMA, SCORE_SCHEMA

import json
import functools

app = Flask(__name__,static_folder=r'templates\assets')
app.json_encoder = CustomJSONEncoder
app.config['DEBUG']= True,
app.config['SESSION_PERMANENT']= True
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_USE_SIGNER'] = False
app.config['TRAP_HTTP_EXCEPTIONS']= True
app.config['PERMANENT_SESSION_LIFETIME']= timedelta(hours=2)
app.config['meeting'] = None
Session(app)
Mobility(app)

config:dict = json.load(open('config.json','r'))

@app.before_first_request
def initialize():
    db.connect()
    db.create_tables(_TABLES_)

@app.before_request
def _before_request():
    if 'db' not in g:
        g.db=db
    
    if not g.db.is_connection_usable():
        g.db.connect()
        
    if not 'meeting' in g:
        g.meeting = Meeting.get_or_none(Meeting.in_progress==True)
    
    if session.get('mac') is None:
        if request.remote_addr in ('localhost', '127.0.0.1'):
            session['local_user'] = True
            if app.debug:
                mac = 'local'
            else:
                return redirect(url_for('admin'))
        else:
            mac = get_mac_address(ip=request.remote_addr)
        device,created = Device.get_or_create(mac=mac)
        if created:
            device.save()
        session['mac'] = mac
        session['device'] = device

@app.teardown_request
def _db_close(exc):
   if not g.db.is_closed():
      g.db.close()
      
@app.errorhandler(400)
def bad_request(error):
    if isinstance(error.description, ValidationError):
        original_error = error.description
        return jsonify({'info': original_error.message}), 400
    return error

@app.route('/')
def index():
    return render_template('students.html',registered = (session['device'].student is not None))

EASTER_EGG = " EASTER EGG: I'm so happy that you are reading this! good luck and hack the planet! BSimjoo ;-)"      # An easter-egg for my students!

@app.route('/api/v1/register')
def register_device():
    device = session['device']
    if (student:=device.student) is None:   #device is not registered
        if not (std_num:=request.args.get('std_num')) in (None,''):
            if (student:=Student.get_or_none(Student.number==std_num)) is not None:
                device.student = student
                device.save()
                session['student'] = student
                return jsonify(name=student.name)
            else:   # student not existed
                return jsonify(info='student not existed. check std_num.'+EASTER_EGG), 404
        else:   # std_num didn't send
            return jsonify(info="std_num didn't send."+EASTER_EGG), 400
    else:   # The student has already registered this device
        return jsonify(name=student.name, info="device is already register, new registration is forbidden."+EASTER_EGG), 403

@app.route('/api/v1/whoami')
def whoami():
    device = session['device']
    if (student:=device.student) is not None:
        return jsonify(name = student.name, number=student.number)
    abort(400)

@app.route('/api/v1/attendance')
def attendance():
    if g.meeting is not None:
        if (device:=session.get('device')) is not None:
            student = device.student
            query = (g.meeting.attendances
                     .select()
                     .join(Student, on=Attendance.student==Student.id)      # type: ignore
                     .where(Student.id==student.id)                          # type: ignore
                    )
            code = 200
            res={}
            if query.count()!=0:
                code = 203
                res['info']='Your presence is already registered.'+EASTER_EGG
            else:
                Attendance.create(student = student, device = device, meeting = g.meeting).save()
            res={
                'name':student.name,
                'number': student.number,
                'attendances': list(student.attendances.dicts()),
                'scores': list(student.scores.dicts()),
                'meetings': list(Meeting.select().dicts()),
                'devices': list(d.to_dict(recurse=False) for d in student.devices),
                'current_meeting': g.meeting.id,
                'total_score': student.total_score,
                'total_full_score': student.total_full_score
            }
            return jsonify(res), code
        else:   # user has not registered yet
            return jsonify(info="You must register first."+EASTER_EGG), 403
    else:
        return jsonify( info="session did not started yet."+EASTER_EGG), 404
    
@app.route('/admin')
def admin():
    if config.get('admin from localhost',True):
        if request.remote_addr not in ['localhost','127.0.0.1']:
            return redirect('/')
    return render_template('admin.html',admin=session.get('admin',False))
    
@app.route('/api/v1/login', methods=['POST'])
@expects_json(LOGIN_SCHEMA)
def login():
    if config.get('admin from localhost',True):
        if request.remote_addr not in ['localhost','127.0.0.1']:
            return redirect('/')
    if session.get('admin'):
        return jsonify(logged_in=True)
    if session.get('tries_left',5) > 0:
        username = g.data['username']
        password = g.data['password']
        if username and password:
            if (config['admin username'],config['admin password']) == (username, password):
                session['admin'] = True
                session['tries_left'] = 5
                return jsonify(logged_in=True)
        session['tries_left'] = session.get('tries_left',5) - 1
        if session['tries_left'] == 0:
            return jsonify(info="login failed, you got banned."+EASTER_EGG), 403
        return jsonify(tries_left=session['tries_left'], info="login failed."+EASTER_EGG), 401
    else:
        return jsonify( info="you are banned."+EASTER_EGG), 403
    
@app.route('/api/v1/can_login')
def can_login():
    if config.get('admin from localhost',True):
        if request.remote_addr not in ['localhost','127.0.0.1']:
            return jsonify(can_login=False, banned = session.get('banned',False)), 403
    if session.get('banned',False):
        return jsonify(can_login=True, banned = True), 403
    return jsonify(can_login=True, banned = False)
    
def login_required(func):
    @functools.wraps(func)
    def wrapper(*args, **kw):
        if config.get('admin from localhost',True):
            if request.remote_addr not in ['localhost','127.0.0.1']:
                return redirect('/')
        if app.debug or session.get('admin'):
            return func(*args,**kw)
        return redirect(url_for('login'))
    return wrapper

@app.route('/api/v1/students',methods = ['POST'])
@login_required
def set_students():
    if (file:=request.files.get('file')):
        if file.filename:
            if file.filename.rsplit('.',1)[1]=='xlsx':
                with NamedTemporaryFile('r+b') as tmp:
                    file.save(tmp)
                    tmp.seek(0)
                    wb = load_workbook(tmp.name)
                    ws = wb.active
                    for row in ws.iter_rows(max_col=2):
                        std = Student(number = row[0].value, name = row[1].value)
                        std.save()
                    return redirect(url_for('admin'))
            else:
                return jsonify( info="bad file format")
    abort(400)              

@app.route('/api/v1/students')
@login_required
def get_students():
    return jsonify([std.to_dict(max_depth=1) for std in Student.select()])
    
@app.route('/api/v1/students/<int:number>')
def get_student(number):
    if session.get('admin') or ((std:=session.get('student')) and std.id==number):
        if (student:=Student.get_or_none(Student.id==number)) is not None:   # type: ignore
            return jsonify(student.to_dict())       # type: ignore
        abort(404)
    return jsonify(info="You're not authorized, get your info or login as admin!"+EASTER_EGG), 401
      
@app.route('/api/v1/attendances')
@login_required
def get_attendances():
    return jsonify([a.to_dict() for a in Attendance.select()])

@app.route('/api/v1/attendances/<int:attendance_id>')
def get_attendance(attendance_id):
    if session.get('admin') or ((std:=session.get('student')) and std.attendances.select(Attendance.id==attendance_id).count() == 1):   #type: ignore
        if (a:=Attendance.get_or_none(Attendance.id==attendance_id)) is not None:   #type: ignore
            return jsonify(a.to_dict())
        abort(404)
    return jsonify(info="You're not authorized, get your attendance info or login as admin!"+EASTER_EGG), 401
    
@app.route('/api/v1/devices')
@login_required
def get_devices():
    return jsonify([ device.to_dict(backrefs=True)for device in Device.select() ])

@app.route('/api/v1/devices/<int:device_id>')
def get_device(device_id):
    if session.get('admin') or ((std:=session.get('student')) and std.devices.select(Device.id==device_id).count() == 1):               #type: ignore
        if (device:=Device.get_or_none(Device.id==device_id)) is not None:      # type: ignore
            return jsonify(device.to_dict(backrefs=True))
        abort(404)
    return jsonify(info="You're not authorized, get your device or login as admin!"+EASTER_EGG), 401
    
@app.route('/api/v1/current_meeting')
@login_required
def get_current_meeting():
    if g.meeting is not None and g.meeting.in_progress:
        return jsonify(g.meeting.to_dict())
    return jsonify(info="no in progress meeting"), 404
    
@app.route('/api/v1/current_meeting',methods=['POST'])
@login_required
def start_meeting():
    if g.meeting is None or not g.meeting.in_progress:
        g.meeting = Meeting.create()
    else:
        return jsonify(info="a meeting is already in progress"), 202
    if g.meeting.save() == 1:
        return jsonify(g.meeting.to_dict())
    return jsonify(done=False,  info="Unknown error while creating database record"), 500
    
@app.route('/api/v1/current_meeting', methods=['DEL'])
@login_required
def end_current_meeting():
    if g.meeting is not None and g.meeting.in_progress:
        g.meeting.in_progress=False
        g.meeting.end_at = datetime.now().time()
        g.meeting.save()
        return jsonify(g.meeting.to_dict())
    return jsonify(info="no in progress meeting"), 404
            
@app.route('/api/v1/meetings')
@login_required
def get_meetings():
    return jsonify([meeting.to_dict(backrefs=False) for meeting in Meeting.select()])

@app.route('/api/v1/meetings/<int:meeting_id>')
@login_required
def get_meeting(meeting_id:int):
    if (meet:=Meeting.get_or_none(Meeting.id==meeting_id)) is not None: #type: ignore
        return jsonify(meet.to_dict(backrefs=True))
    abort(404)
    
@app.route('/api/v1/meetings/<int:meeting_id>',methods=['DEL'])
@login_required
def del_meetings(meeting_id:int):
    if (meet:=Meeting.get_or_none(Meeting.id==meeting_id)) is not None: #type: ignore
        if (rows:=meet.delete_instance(True)) >0:
            if (meet==g.meeting):
                g.meeting = None
            return jsonify(rows=rows)
        return jsonify(info="didn't deleted any row"), 500
    abort(404)
    
@app.route('/api/v1/score', methods=["POST"])
@login_required
@expects_json(SCORE_SCHEMA)
def add_edit_score():
    try:
        student = Student.get_by_id(g.data['student'])
        meeting = g.data['meeting'] and Meeting.get_by_id(g.data['meeting'])
        score = g.data['id'] and Score.get_by_id(g.data['id'])
    except DoesNotExist:
        abort(404)
    if score:
        score.score = g.data['score']
        score.full_score = g.data['full_score']
        score.reason = g.data['reason']
    else:
        score = Score.create(
            student = student,
            score = g.data['score'],
            full_score = g.data['full_score'],
            meeting = meeting,
            reason = g.data['reason']
        )
    res = score.save()
    return jsonify(done=res==1), 200 if res==1 else 500
