from openpyxl import load_workbook
from model import (  # noqa:F401
    database_proxy,
    Student,
    Device,
    Attendance,
    Score,
    Meeting,
    _TABLES_,
)
from playhouse.db_url import connect
from typing import Callable
import json
import re
import argparse

config: dict = json.load(open("config.json", "r"))

try:
    db = connect(config["database"])
    database_proxy.initialize(db)
except:  # noqa: E722
    print("[ERROR] can not connect to database, are you sure you shuted server down?!")
    exit(99)


def get_input(
    condition: Callable,
    *prompt,
    sep=" ",
    error: list | tuple | str = "[ERROR} Bad input",
):
    while True:
        inp = input(sep.join(prompt))
        response = condition(inp)
        result = None
        if isinstance(response, tuple) and len(response) == 2:
            response, result = response

        if (isinstance(response, bool) and response) or (
            isinstance(response, int) and response == 0
        ):
            if result:
                return result
            else:
                return inp
        elif isinstance(response, int) and isinstance(error, (list, tuple)):
            print(error[response - 1])
        else:
            print(error)


validate_range_input = re.compile(r"([A-Z]+)(\d+):((\1\d+)|([A-Z]+)\2)")


def validate_array(inp, ws, len_match=0):
    if validate_range_input.fullmatch(inp) is None:
        return 1
    start, end = inp.split(":")
    array = ws[start:end]
    if len(array) == 1 and len(array[0]) > 1:
        array = [cell.value for cell in array[0]]
    elif len(array) > 1 and len(array[0]) == 1:
        array = [cell[0].value for cell in array]
    else:
        return 2

    array = list(filter(None, array))
    if not array:
        return 3
    if len_match > 0 and len(array) != len_match:
        return 4
    return 0, array


def menu(prompt, *options):
    print(prompt)
    for i, option in enumerate(options, 1):
        print(f"  {i}- {option}")

    while True:
        inp = input(f"Enter [1-{len(options)}]: ")
        if inp.isdecimal() and 1 < int(inp) <= len(options):
            return int(inp)
        else:
            print("Bad input.", end=" ")


def load(file):
    if any(map(lambda table: table.table_exists(), _TABLES_)):
        if input("Do you really wish to overwrite? (yes/no)").lower() != "yes":
            return 1

        for table in _TABLES_:
            table.drop_table()
    db.create_tables(_TABLES_)
    wb = load_workbook(file)
    if len(wb.sheetnames) > 1:
        index = menu("Choose a worksheet:", *wb.sheetnames)
        ws = wb.worksheets[index]
    else:
        ws = wb.active

    print(
        f'opened "{ws.title}" now please give address of data, you can use excel to find ranges, "\
            "students name and number must match one-to-one, means first student name saves by first student number.'
    )

    students_name = get_input(
        lambda inp: validate_array(inp, ws),
        "where are the students name? (e,g: A1:A12) : ",
        error=[
            "[ERROR] Range is invalid",
            "[ERROR] Are you kidding me?! you have just one student?!",
            "[ERROR] really?! you haven't any student?! I don't think so!",
        ],
    )

    students_numbers = get_input(
        lambda inp: validate_array(inp, ws, len(students_name)),
        "and where are the students number?(B1:B12) : ",
        error=[
            "[ERROR] Range is invalid",
            "[ERROR] Are you kidding me?! just one student number?!",
            "[ERROR] Did't found any value",
            "[ERROR] numbers and names count mismatch",
        ],
    )

    for i, (name, number) in enumerate(zip(students_name, students_numbers), 1):
        print(f"[{i}] add {name} with number {number}")
        std = Student.create(name=name, number=number)
        res = std.save()
        if res != 1:
            print(f"[Error] Something went wrong, database returned {res} while saving")
            return 2
    print("---[Congratulation, All done]---")
    return 0


def add(student_name, student_number):
    db.create_tables(_TABLES_)
    if Student.get_or_none(Student.number == student_number):
        print(f'[ERROR] A student with number "{student_number}" already existed.')
        return 1

    if Student.get_or_none(Student.name == student_name):
        print(f'[ERROR] A student with name "{student_name}" already existed.')
        return 2

    std = Student.create(name=student_name, number=student_number)
    if res := (std.save()) != 1:
        print(f"[Error] Something went wrong, database returned {res} while saving.")
        return 3
    print(
        f'Successfully added new student (id:{std.id}, name:"{std.name}", number:{std.number})'
    )
    return 0


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        "studmgr.py", description="This script will import your excel worksheet."
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument(
        "--load",
        "-l",
        nargs=1,
        metavar='"worksheet file path"',
        type=argparse.FileType("rb"),
        help="loads a excel worksheet file and imports it to database",
    )
    group.add_argument(
        "--add",
        "-a",
        nargs=2,
        metavar=("name", "number"),
        help="manually add a new student to database",
    )
    args = parser.parse_args()
    if args.load is not None:
        exit(load(args.load))
    if args.add is not None:
        exit(add(*args.add))
